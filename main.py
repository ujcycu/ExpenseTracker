import os
import sqlite3
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Optional
from fastapi import FastAPI, HTTPException, status, UploadFile, File
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from fastapi.middleware.cors import CORSMiddleware
from datetime import date, timedelta

# 定義資料庫檔案名稱
DB_FILE = "accounting.db"

app = FastAPI(
    title="高效防呆記帳系統 API (升級版)",
    description="支援動態母子錢包、錢包封存、類 Excel 批次編輯與 Excel 匯入匯出對帳核心",
    version="1.1.0"
)

# 避免觸發 CORS 跨網域封鎖
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # 允許所有前端網頁造訪
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# 1. Pydantic 資料驗證模型 (前端傳入規格與防呆)
# ==========================================

class BulkRecordItem(BaseModel):
    id: Optional[int] = Field(None, description="紀錄 ID。新列填 None，舊資料修改填原本的 ID")
    date: str = Field(default_factory=lambda: date.today().isoformat(), description="交易日期 YYYY-MM-DD")
    type: str = Field(..., pattern="^(income|expense|transfer)$", description="交易類型")
    amount: int = Field(..., gt=0, description="防呆：金額必須大於 0")
    category: str = Field(..., min_length=1, description="分類不能為空字串")
    note: Optional[str] = None
    from_wallet_id: Optional[int] = None
    to_wallet_id: Optional[int] = None
    recon_status: str = Field("unreconciled", pattern="^(unreconciled|matched|billed)$")

class WalletCreateItem(BaseModel):
    name: str = Field(..., min_length=1, description="錢包名稱不能為空")
    type: str = Field(..., pattern="^(cash|bank|credit_card|loan)$", description="錢包類型")
    parent_id: Optional[int] = Field(None, description="所屬母錢包 ID (若為子錢包則填寫)")

class WalletUpdateParentItem(BaseModel):
    parent_id: Optional[int] = Field(None, description="新的母錢包 ID，解除綁定填 None")


# ==========================================
# 2. 資料庫核心工具與初始化
# ==========================================

def get_db_connection():
    """建立資料庫連線，並強制啟用 SQLite 的外鍵約束"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row  # 讓查詢結果可以用欄位名稱讀取
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

@app.on_event("startup")
def init_db():
    """系統啟動時，自動建立或更新結構健全的資料表"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 建立錢包表 (升級版：支援 parent_id 實作母子架構、is_archived 實作封存)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS wallets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        type TEXT NOT NULL,
        balance INTEGER NOT NULL DEFAULT 0,
        parent_id INTEGER DEFAULT NULL,
        is_archived INTEGER NOT NULL DEFAULT 0, -- 0: 正常使用, 1: 已封存
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (parent_id) REFERENCES wallets(id) ON DELETE SET NULL,
        CONSTRAINT CHK_WalletType CHECK (type IN ('cash', 'bank', 'credit_card', 'loan')),
        CONSTRAINT CHK_IsArchived CHECK (is_archived IN (0, 1))
    );
    """)
    
    # 建立帳單表
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS statements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        wallet_id INTEGER NOT NULL,
        statement_period TEXT NOT NULL,
        bank_total_amount INTEGER NOT NULL,
        is_paid INTEGER NOT NULL DEFAULT 1,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (wallet_id) REFERENCES wallets(id)
    );
    """)
    
    # 建立紀錄表
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        type TEXT NOT NULL,
        amount INTEGER NOT NULL,
        category TEXT NOT NULL,
        note TEXT,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        from_wallet_id INTEGER,
        to_wallet_id INTEGER,
        recon_status TEXT NOT NULL DEFAULT 'unreconciled',
        statement_id INTEGER,
        FOREIGN KEY (from_wallet_id) REFERENCES wallets(id),
        FOREIGN KEY (to_wallet_id) REFERENCES wallets(id),
        FOREIGN KEY (statement_id) REFERENCES statements(id),
        CONSTRAINT CHK_AmountPositive CHECK (amount > 0),
        CONSTRAINT CHK_ValidType CHECK (type IN ('income', 'expense', 'transfer')),
        CONSTRAINT CHK_ReconStatus CHECK (recon_status IN ('unreconciled', 'matched', 'billed'))
    );
    """)
    
    # 💡 塞入富有「母子錢包情境」的初始測試資料 (Seed Data) 如果表是空的
    cursor.execute("SELECT COUNT(*) FROM wallets;")
    if cursor.fetchone()[0] == 0:
        # 先建立虛擬的母錢包群組 (類型通常跟著子錢包，這裡可用 bank 或建一個通用概念)
        cursor.execute("INSERT INTO wallets (name, type, balance) VALUES (?, ?, ?);", ("KGI", "bank", 0))
        fubon_parent_id = cursor.lastrowid
        
        cursor.execute("INSERT INTO wallets (name, type, balance) VALUES (?, ?, ?);", ("生活開銷群組", "cash", 0))
        life_parent_id = cursor.lastrowid

        # 塞入實際的子錢包與獨立錢包
        cursor.executemany("""
            INSERT INTO wallets (name, type, balance, parent_id) VALUES (?, ?, ?, ?);
        """, [
            ("交割", "bank", 0, fubon_parent_id),
            ("STOCK", "bank", 0, fubon_parent_id)
        
        ])
            # ("個人皮夾", "cash", 4766, life_parent_id),
            # ("iPass Money", "cash", 0, life_parent_id),
            # ("國泰信用卡", "credit_card", 0, None),  # 獨立錢包，無母錢包
            # ("保險分期", "loan", -40000, None)
    
    conn.commit()
    conn.close()


# ==========================================
# 3. 錢包與母子看板管理 API 端點
# ==========================================

@app.get("/api/wallets", summary="獲取所有未封存錢包明細、母子帳務聚合看板、及系統頂部資產總計")
def get_wallets():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 💡 核心修正：不再直接讀取 w.balance，而是用 SQL 即時去 records 明細表算出來！
    # 錢包最新餘額 = 初始餘額 (w.balance) + 流入的總額 (to_wallet) - 流出的總額 (from_wallet)
    # 註：如果你的錢包初始餘額欄位叫 initial_balance，請把下面的 w.balance 改成 w.initial_balance
    sql_query = """
    SELECT 
        w.id, 
        w.name, 
        w.type, 
        w.parent_id,
        w.balance
    FROM wallets w 
    WHERE w.is_archived = 0;
    """
    
    cursor.execute(sql_query)
    rows = cursor.fetchall()
    conn.close()
    
    all_wallets = [dict(row) for row in rows]
    
    # 建立索引對照
    wallet_dict = {w["id"]: w for w in all_wallets}
    
    # 動態計算：初始化母錢包的加總金額
    parent_balances = {}
    total_sys_balance = 0 # 頂部看板的資產總金額
    
    # 第一輪：先把身為子錢包的金額，加到其母錢包的虛擬計數器中
    for w in all_wallets:
        total_sys_balance += w["balance"] # 累加系統總額
        p_id = w["parent_id"]
        if p_id:
            if p_id not in parent_balances:
                parent_balances[p_id] = 0
            parent_balances[p_id] += w["balance"]
            
    # 將計算出的母錢包聚合總額，塞回資料中方便前端直接渲染
    for w in all_wallets:
        if w["id"] in parent_balances:
            w["aggregated_balance"] = parent_balances[w["id"]]
        else:
            w["aggregated_balance"] = w["balance"] # 若自己是子錢包或獨立錢包，則為自身餘額

    return {
        "total_system_balance": total_sys_balance, # 主畫面上方總額
        "wallets": all_wallets                     # 完整的錢包清單與關係
    }


@app.post("/api/wallets/create", summary="使用者自行新增錢包（可直接指定母錢包）")
def create_wallet(item: WalletCreateItem):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO wallets (name, type, balance, parent_id) 
            VALUES (?, ?, 0, ?);
        """, (item.name, item.type, item.parent_id))
        conn.commit()
        return {"status": "success", "wallet_id": cursor.lastrowid, "message": f"成功新增錢包：{item.name}"}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="錢包名稱重複或母錢包 ID 不存在")
    finally:
        conn.close()


@app.put("/api/wallets/{wallet_id}/archive", summary="封存錢包（隱藏且不計入主畫面總額）")
def archive_wallet(wallet_id: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 防呆：檢查是否存在
    cursor.execute("SELECT name FROM wallets WHERE id = ?;", (wallet_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="找不到該錢包")
        
    cursor.execute("UPDATE wallets SET is_archived = 1 WHERE id = ?;", (wallet_id,))
    conn.commit()
    conn.close()
    return {"status": "success", "message": f"錢包 [{row['name']}] 已成功封存"}


@app.put("/api/wallets/{wallet_id}/unarchive", summary="解封錢包")
def unarchive_wallet(wallet_id: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE wallets SET is_archived = 0 WHERE id = ?;", (wallet_id,))
    conn.commit()
    conn.close()
    return {"status": "success", "message": "錢包已成功解除封存"}


@app.put("/api/wallets/{wallet_id}/update-parent", summary="動態調整母子關係（將子錢包移入或移出某母錢包）")
def update_wallet_parent(wallet_id: int, item: WalletUpdateParentItem):
    if wallet_id == item.parent_id:
        raise HTTPException(status_code=400, detail="防呆：母錢包不能設為自己本身")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE wallets SET parent_id = ? WHERE id = ?;", (item.parent_id, wallet_id))
        conn.commit()
        return {"status": "success", "message": "母子錢包對應關係調整成功"}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="指定的母錢包 ID 不存在")
    finally:
        conn.close()


# ==========================================
# 4. 記帳明細 API 端點
# ==========================================

@app.get("/api/records", summary="篩選日常編輯主表格明細（排除已歸檔封存資料）")
def get_records(
    days: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    wallet_id: Optional[int] = None
):
    conn = get_db_connection()
    cursor = conn.cursor()

    # 基本條件：排除已月底歸檔的
    query = "SELECT id, date, type, amount, category, note, from_wallet_id, to_wallet_id, recon_status "
    query += " FROM records WHERE recon_status != 'billed'"
    params = []
    
    # 篩選條件 A：N天之內
    if days is not None:
        target_date = (date.today() - timedelta(days=days)).isoformat()
        query += " AND date >= ?"
        params.append(target_date)
    
    # 篩選條件 B：特定日期區間
    if start_date:
        query += " AND date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
        
    # 篩選條件 C：檢視某個特定錢包的交易資料
    if wallet_id is not None:
        query += " AND (from_wallet_id = ? OR to_wallet_id = ?)"
        params.append(wallet_id)
        params.append(wallet_id)
        
    # 排序
    query += " ORDER BY date ASC, id ASC;"
    
    cursor.execute(query, params)
    records = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return records


@app.post("/api/records/bulk-save", summary="【核心防呆】類 Excel 批次原子化儲存端點")
def bulk_save_records(items: List[BulkRecordItem]):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        conn.execute("BEGIN TRANSACTION;")

        # 先撈出所有錢包的型態對照表，方便在迴圈中快速判斷
        cursor.execute("SELECT id, type FROM wallets;")
        wallet_types = {row["id"]: row["type"] for row in cursor.fetchall()}
        
        for index, item in enumerate(items):
            # === 【核心防禦：防重複扣款】 ===
            if item.id is not None:
                # 1. 撈出該 ID 存在資料庫的「舊狀態」
                cursor.execute("SELECT type, amount, from_wallet_id, to_wallet_id, recon_status FROM records WHERE id = ?;", (item.id,))
                old = cursor.fetchone()
                
                if old:
                    old_from_type = wallet_types.get(old["from_wallet_id"])
                    old_to_type = wallet_types.get(old["to_wallet_id"])
                    
                    # 💡 【舊 From 錢包回滾】
                    if old["from_wallet_id"]:
                        # 現金/銀行是在日常(不論狀態)就扣款，所以一律加回來
                        if old_from_type in ('cash', 'bank', 'loan'):
                            cursor.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?;", (old["amount"], old["from_wallet_id"]))
                        # 信用卡只有在已對帳(matched)時才算扣款，所以舊狀態是 matched 才需要加回來
                        elif old_from_type == 'credit_card' and old["recon_status"] == 'matched':
                            cursor.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?;", (old["amount"], old["from_wallet_id"]))

                    # 💡 【舊 To 錢包回滾】
                    if old["to_wallet_id"]:
                        if old_to_type in ('cash', 'bank', 'loan'):
                            cursor.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?;", (old["amount"], old["to_wallet_id"]))
                        elif old_to_type == 'credit_card' and old["recon_status"] == 'matched':
                            cursor.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?;", (old["amount"], old["to_wallet_id"]))
            
            # === 2. 驗證新欄位規則 ===
            if item.type == "expense" and item.from_wallet_id is None:
                raise HTTPException(status_code=400, detail=f"第 {index + 1} 行：支出缺少 From 錢包")
            if item.type == "income" and item.to_wallet_id is None:
                raise HTTPException(status_code=400, detail=f"第 {index + 1} 行：收入缺少 To 錢包")
            if item.type == "transfer" and (item.from_wallet_id is None or item.to_wallet_id is None):
                raise HTTPException(status_code=400, detail=f"第 {index + 1} 行：轉帳缺少錢包")

            # === 3. 寫入新資料 / 更新舊資料 ===
            if item.id is None:
                cursor.execute("""
                    INSERT INTO records (date, type, amount, category, note, from_wallet_id, to_wallet_id, recon_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?);
                """, (item.date, item.type, item.amount, item.category, item.note, item.from_wallet_id, item.to_wallet_id, item.recon_status))
            else:
                cursor.execute("""
                    UPDATE records SET date=?, type=?, amount=?, category=?, note=?, from_wallet_id=?, to_wallet_id=?, recon_status=? WHERE id=?;
                """, (item.date, item.type, item.amount, item.category, item.note, item.from_wallet_id, item.to_wallet_id, item.recon_status, item.id))
            
            # === 4. 套用新餘額扣減 ===
            new_from_type = wallet_types.get(item.from_wallet_id)
            new_to_type = wallet_types.get(item.to_wallet_id)
            
            # 💡 【新 From 錢包扣款】
            if item.from_wallet_id:
                if new_from_type in ('cash', 'bank', 'loan'):
                    # 現金/銀行/貸款：日常即扣款
                    cursor.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?;", (item.amount, item.from_wallet_id))
                elif new_from_type == 'credit_card' and item.recon_status == 'matched':
                    # 信用卡：只有在勾選「已核對」時才扣款（增加負債）
                    cursor.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?;", (item.amount, item.from_wallet_id))

            # 💡 【新 To 錢包流入】
            if item.to_wallet_id:
                if new_to_type in ('cash', 'bank', 'loan'):
                    # 現金/銀行/貸款：日常即增加入帳
                    cursor.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?;", (item.amount, item.to_wallet_id))
                elif new_to_type == 'credit_card' and item.recon_status == 'matched':
                    # 信用卡：只有在勾選「已核對」時才影響金額（例如退刷或還款）
                    cursor.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?;", (item.amount, item.to_wallet_id))

        conn.commit()
        return {"status": "success", "message": "批次處理成功（已自動計算回滾防重複扣款）"}
    except Exception as e:
        conn.rollback()
        raise e if isinstance(e, HTTPException) else HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ==========================================
# 5. Excel (CSV 格式) 匯入/匯出與批量測試端點
# ==========================================

@app.get("/api/records/export-excel", summary="將現有的未對帳明細匯出為真正的 Excel (.xlsx) 檔案")
def export_excel_records():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 取得錢包 ID 到名字的對照表
    cursor.execute("SELECT id, name FROM wallets;")
    wallet_names = {row["id"]: row["name"] for row in cursor.fetchall()}
    
    cursor.execute("SELECT id, date, type, amount, category, note, from_wallet_id, to_wallet_id, recon_status FROM records WHERE recon_status != 'billed';")
    records = cursor.fetchall()
    conn.close()
    
    # 建立 Excel 活頁簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "未對帳明細"
    ws.views.sheetView[0].showGridLines = True # 確保格線看得到
    
    # 1. 寫入標頭
    headers = ["紀錄ID(系統使用，新增免填)", "交易日期(YYYY-MM-DD)", "交易類型(income/expense/transfer)", "金額", "分類", "備註", "從哪個錢包(填名稱)", "到哪個錢包(填名稱)", "核對狀態"]
    ws.append(headers)
    
    # 2. 寫入資料
    for r in records:
        ws.append([
            r["id"],
            r["date"],
            r["type"],
            r["amount"],
            r["category"],
            r["note"] or "",
            wallet_names.get(r["from_wallet_id"], ""),
            wallet_names.get(r["to_wallet_id"], ""),
            r["recon_status"]
        ])
        
    # 3. 美化樣式 (深藍色質感知性風)
    font_family = "Microsoft JhengHei"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    body_font = Font(name=font_family, size=10, bold=False, color="000000")
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_style = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 格式化標頭
    ws.row_dimensions[1].height = 28
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_style
        
    # 格式化資料列
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(records) + 1), 2):
        ws.row_dimensions[row_idx].height = 22
        is_even = (row_idx % 2 == 0)
        
        for col_idx, cell in enumerate(row, 1):
            cell.font = body_font
            cell.border = border_style
            if is_even:
                cell.fill = zebra_fill
                
            # 依欄位調整對齊與千分位格式
            if col_idx in [1, 4]:  # ID & 金額
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_idx == 4:
                    cell.number_format = '#,##0'
            elif col_idx == 2:     # 日期
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # 4. 自動適應調整欄寬
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                cell_len = sum(2 if ord(char) > 127 else 1 for char in val_str)
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # 5. 打包二進位流回傳
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=accounting_records.xlsx"}
    )


@app.post("/api/records/import-excel", summary="完整版：批次匯入 Excel 紀錄並自動連動錢包餘額")
def import_excel_records(file: UploadFile = File(...)):
    # 1. 檢查副檔名
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="請上傳標準的 Excel (.xlsx) 檔案。")
        
    try:
        # 2. 讀取 Excel 檔案內容
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        raw_rows = []
        for row in ws.iter_rows(values_only=True):
            # 略過完全空白的無效行
            if any(cell is not None for cell in row):
                raw_rows.append(row)
                
        if not raw_rows or len(raw_rows) < 2:
            raise HTTPException(status_code=400, detail="Excel 內沒有有效的資料紀錄（至少需包含標題列與一筆資料）。")
            
        header = raw_rows[0]
        data_rows = raw_rows[1:]
        
        # 3. 連線資料庫並預先載入「錢包名稱 -> ID」對照表
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row  # 確保可以透過名稱存取欄位
        cursor = conn.cursor()
        
        cursor.execute("SELECT id, name FROM wallets;")
        wallet_name_to_id = {row["name"]: row["id"] for row in cursor.fetchall()}
        
        success_count = 0
        error_logs = []
        
        print("\n" + "="*50)
        print(" 🔍 [Excel 匯入除錯監控] 開始逐行解析...")
        print(f" 當前系統擁有的錢包清單: {list(wallet_name_to_id.keys())}")
        print("="*50)
        
        # 4. 逐行解析 Excel
        # 假設你的欄位順序為：
        # idx 0:紀錄ID, idx 1:交易日期, idx 2:交易類型, idx 3:金額, idx 4:分類, idx 5:備註, idx 6:從哪個錢包, idx 7:到哪個錢包, idx 8:核對狀態
        for line_num, row in enumerate(data_rows, start=2):
            print(f"\n👉 [第 {line_num} 行原始資料]: {row}")
            
            # 安全防呆：如果這行長度不夠，直接跳過或補齊
            if len(row) < 8:
                print(f"  ❌ 錯誤：第 {line_num} 行欄位數量不足，跳過該行。")
                error_logs.append(f"第 {line_num} 行：欄位數量不足。")
                continue
                
            try:
                # 擷取並去空白
                record_date = str(row[1]).strip() if row[1] else None
                record_type = str(row[2]).strip().lower() if row[2] else None  # 強制轉小寫相容 income/expense/transfer
                amount_raw = row[3]
                category = str(row[4]).strip() if row[4] else ""
                note = str(row[5]).strip() if row[5] else ""
                from_wallet_name = str(row[6]).strip() if row[6] else ""
                to_wallet_name = str(row[7]).strip() if row[7] else ""
                recon_status = str(row[8]).strip() if len(row) > 8 and row[8] else "reconciled"
                
                # 數據基本校驗
                if not record_date or not record_type:
                    print(f"  ❌ 錯誤：日期或交易類型為空，跳過該行。")
                    error_logs.append(f"第 {line_num} 行：日期或類型不可為空。")
                    continue
                    
                # 轉型金額
                try:
                    amount = int(float(amount_raw)) if amount_raw is not None else 0
                except ValueError:
                    print(f"  ❌ 錯誤：金額 '{amount_raw}' 無法轉換為數字，跳過該行。")
                    error_logs.append(f"第 {line_num} 行：金額格式錯誤。")
                    continue
                
                # 比對錢包名稱並取得對應 ID
                from_wallet_id = wallet_name_to_id.get(from_wallet_name) if from_wallet_name else None
                to_wallet_id = wallet_name_to_id.get(to_wallet_name) if to_wallet_name else None
                
                # 偵錯印出名稱比對結果
                print(f"  -> 解析結果: 類型={record_type}, 金額={amount}, 分類={category}")
                print(f"  -> 錢包對照: '{from_wallet_name}'(ID:{from_wallet_id}) ➡️ '{to_wallet_name}'(ID:{to_wallet_id})")
                
                # 核心防呆：如果欄位有寫名稱，但系統找不到這個錢包 ID，提示使用者
                if from_wallet_name and from_wallet_id is None:
                    print(f"  ⚠️ 警告：系統內找不到名為 '{from_wallet_name}' 的錢包！")
                    raise HTTPException(
                        status_code=400, 
                        detail=f"匯入失敗！第 {line_num} 行的來源錢包名稱「{from_wallet_name}」在系統中不存在，請修正名稱後重新上傳。"
                    )
                if to_wallet_name and to_wallet_id is None:
                    print(f"  ⚠️ 警告：系統內找不到名為 '{to_wallet_name}' 的錢包！")
                    raise HTTPException(
                        status_code=400, 
                        detail=f"匯入失敗！第 {line_num} 行的目標錢包名稱「{to_wallet_name}」在系統中不存在，請修正名稱後重新上傳。"
                    )
                
                # 交易類型連動錢包的防呆
                if record_type == "income" and not to_wallet_id:
                    raise HTTPException(status_code=400, detail=f"第 {line_num} 行：交易類型為收入(income)，但「到哪個錢包」不可以是空白。")
                if record_type == "expense" and not from_wallet_id:
                    raise HTTPException(status_code=400, detail=f"第 {line_num} 行：交易類型為支出(expense)，但「從哪個錢包」不可以是空白。")
                if record_type == "transfer" and (not from_wallet_id or not to_wallet_id):
                    raise HTTPException(status_code=400, detail=f"第 {line_num} 行：交易類型為轉帳(transfer)，「從」與「到」錢包皆不可為空白。")
                
                # 5. 寫入 records 明細表
                cursor.execute("""
                    INSERT INTO records (date, type, amount, category, note, from_wallet_id, to_wallet_id, recon_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?);
                """, (record_date, record_type, amount, category, note, from_wallet_id, to_wallet_id, recon_status))
                
                # 6. 連動更新 wallets 表的 balance 欄位
                if record_type == "income" and to_wallet_id:
                    cursor.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?;", (amount, to_wallet_id))
                    print(f"  💰 錢包連動：幫錢包ID {to_wallet_id} 成功增加 {amount} 元")
                    
                elif record_type == "expense" and from_wallet_id:
                    cursor.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?;", (amount, from_wallet_id))
                    print(f"  💸 錢包連動：幫錢包ID {from_wallet_id} 成功扣除 {amount} 元")
                    
                elif record_type == "transfer":
                    if from_wallet_id:
                        cursor.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?;", (amount, from_wallet_id))
                        print(f"  🔄 轉帳連動：來源錢包ID {from_wallet_id} 扣除 {amount} 元")
                    if to_wallet_id:
                        cursor.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?;", (amount, to_wallet_id))
                        print(f"  🔄 轉帳連動：目標錢包ID {to_wallet_id} 增加 {amount} 元")
                
                success_count += 1
                print(f"  ✅ 第 {line_num} 行處理成功並已寫入緩存。")
            
            # 遇上我們故意拋出的 HTTPException（如錢包不存在），直接往外丟，中斷整個大迴圈！
            except HTTPException as http_err:
                raise http_err
                
            except Exception as row_err:
                print(f"  ❌ 第 {line_num} 行在處理時發生未預期嚴重錯誤: {row_err}")
                error_logs.append(f"第 {line_num} 行執行失敗: {str(row_err)}")
        
        # 7. 💡 最終 Commit！所有步驟都沒噴錯，才一次寫入硬碟硬化資料
        conn.commit()
        conn.close()
        
        print("\n" + "="*50)
        print(f" 🏁 匯入結束！成功寫入總數: {success_count} 筆。")
        print("="*50 + "\n")
        
        return {
            "status": "success" if success_count > 0 else "failed",
            "message": f"Excel 批次處理完畢。成功：{success_count} 筆，失敗：{len(error_logs)} 筆。",
            "errors": error_logs
        }
        
    except Exception as e:
        # 發生全面性崩潰時的安全回滾機制
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=f"伺服器處理 Excel 失敗: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    # 自動啟動本地伺服器
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)